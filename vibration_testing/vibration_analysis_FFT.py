"""
FFT Vibration Data Analysis Script
Processes CSV files containing accelerometer data and generates FFT (frequency spectrum) graphs
with dominant frequency peaks labeled.

Data Specifications:
- Sampling Rate: 100 Hz
- Raw Input: Acceleration data in g (gravitational acceleration units, 1g = 9.81 m/s²)
- Output: Frequency spectrum showing amplitude vs frequency

Processing Method:
- Applies FFT to acceleration magnitude data
- Identifies dominant frequency peaks
- Displays frequency content from 0 to Nyquist frequency (50 Hz)
"""

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from pathlib import Path
import glob
from scipy import signal

# Configuration
DATA_FOLDER = r"C:\Users\dvier\Documents\Thesis\Vibration Testing"
OUTPUT_FOLDER = r"C:\Users\dvier\Documents\Thesis\Vibration Testing\Output_FFT_Graphs"
SAMPLING_RATE = 100  # Hz
G_TO_MS2 = 9.81  # Conversion factor: 1g = 9.81 m/s²

# FFT parameters
FREQUENCY_RANGE = (0, 50)  # Hz - display range (0 to Nyquist frequency)
NUM_PEAKS_TO_LABEL = 3     # Number of dominant peaks to label on each plot

# Label mapping for accelerometer locations
LABEL_MAPPING = {
    'MPU_Direct': 'Bed Plate',
    'MPU_Ch0': 'Vertical Extrusion',
    'MPU_Ch1': 'Tool Head'
}

# Color scheme for each accelerometer
COLORS = {
    'Bed Plate': '#2E86AB',      # Blue
    'Vertical Extrusion': '#A23B72',  # Purple
    'Tool Head': '#F18F01'        # Orange
}


def compute_fft(signal_data, sampling_rate):
    """
    Compute the FFT of a signal and return frequency and magnitude arrays.
    
    Parameters:
    -----------
    signal_data : array-like
        Input signal (acceleration data)
    sampling_rate : float
        Sampling rate in Hz
    
    Returns:
    --------
    tuple : (frequencies, magnitudes)
        frequencies: Array of frequency bins (Hz)
        magnitudes: Array of magnitude values (amplitude spectrum)
    """
    # Number of samples
    N = len(signal_data)
    
    # Remove DC component (mean)
    signal_data = signal_data - np.mean(signal_data)
    
    # Apply Hanning window to reduce spectral leakage
    window = np.hanning(N)
    signal_windowed = signal_data * window
    
    # Compute FFT
    fft_values = np.fft.rfft(signal_windowed)
    
    # Compute magnitude spectrum (normalized)
    magnitudes = np.abs(fft_values) * 2.0 / N
    
    # Compute frequency bins
    frequencies = np.fft.rfftfreq(N, d=1.0/sampling_rate)
    
    return frequencies, magnitudes


def find_dominant_peaks(frequencies, magnitudes, num_peaks=3, min_frequency=1.0):
    """
    Find the dominant frequency peaks in the spectrum.
    
    Parameters:
    -----------
    frequencies : array-like
        Frequency array (Hz)
    magnitudes : array-like
        Magnitude array
    num_peaks : int
        Number of peaks to find (default: 3)
    min_frequency : float
        Minimum frequency to consider (Hz) - avoids DC and very low frequencies
    
    Returns:
    --------
    list : List of tuples (frequency, magnitude) for dominant peaks
    """
    # Filter out frequencies below minimum
    valid_indices = frequencies >= min_frequency
    valid_freqs = frequencies[valid_indices]
    valid_mags = magnitudes[valid_indices]
    
    # Find peaks using scipy's find_peaks
    # prominence ensures we get significant peaks, not just noise
    peak_indices, properties = signal.find_peaks(valid_mags, 
                                                  prominence=np.max(valid_mags)*0.1,
                                                  distance=5)
    
    if len(peak_indices) == 0:
        return []
    
    # Get peak frequencies and magnitudes
    peak_freqs = valid_freqs[peak_indices]
    peak_mags = valid_mags[peak_indices]
    
    # Sort by magnitude (descending) and take top num_peaks
    sorted_indices = np.argsort(peak_mags)[::-1]
    top_indices = sorted_indices[:min(num_peaks, len(sorted_indices))]
    
    # Create list of (frequency, magnitude) tuples
    dominant_peaks = [(peak_freqs[i], peak_mags[i]) for i in top_indices]
    
    # Sort by frequency for consistent display
    dominant_peaks.sort(key=lambda x: x[0])
    
    return dominant_peaks


def process_csv_file(file_path):
    """
    Process a single CSV file and generate FFT plot.
    
    The CSV file contains accelerometer data with x, y, z components (in g) and 
    magnitude values (also in g) for each accelerometer location.
    This function computes the FFT of the acceleration magnitude for each location.
    
    Parameters:
    -----------
    file_path : str or Path
        Path to the CSV file
    """
    # Read the CSV file
    df = pd.read_csv(file_path)
    
    # Extract filename without extension for the plot title
    filename = Path(file_path).stem
    
    # Create figure and axis
    fig, ax = plt.subplots(figsize=(14, 8))
    
    # Dictionary to store peak information
    peak_info = {}
    
    # Process each accelerometer location
    for old_label, new_label in LABEL_MAPPING.items():
        # Get acceleration magnitude column
        mag_column = f'{old_label}_mag'
        
        if mag_column in df.columns:
            # Get acceleration magnitude (in g)
            acc_mag_g = df[mag_column].values
            
            # Convert g to m/s²
            acc_mag_ms2 = acc_mag_g * G_TO_MS2
            
            # Compute FFT
            frequencies, magnitudes = compute_fft(acc_mag_ms2, SAMPLING_RATE)
            
            # Find dominant peaks
            dominant_peaks = find_dominant_peaks(frequencies, magnitudes, 
                                                 num_peaks=NUM_PEAKS_TO_LABEL)
            
            # Store peak information
            peak_info[new_label] = dominant_peaks
            
            # Plot the FFT spectrum
            ax.plot(frequencies, magnitudes, 
                   label=new_label,
                   color=COLORS[new_label],
                   linewidth=1.5,
                   alpha=0.8)
            
            # Mark dominant peaks with markers only (no text annotations)
            peak_freqs = [p[0] for p in dominant_peaks]
            peak_mags = [p[1] for p in dominant_peaks]
            
            ax.plot(peak_freqs, peak_mags, 
                   marker='o', 
                   markersize=8, 
                   color=COLORS[new_label],
                   markeredgecolor='black',
                   markeredgewidth=1.5,
                   linestyle='None',
                   zorder=5)
    
    # Customize the plot
    ax.set_xlabel('Frequency (Hz)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Magnitude (m/s²)', fontsize=12, fontweight='bold')
    ax.set_title(f'FFT Analysis: {filename}\nFrequency Spectrum (Sampling: {SAMPLING_RATE} Hz)', 
                fontsize=14, fontweight='bold', pad=20)
    
    # Set x-axis limits to specified frequency range
    ax.set_xlim(FREQUENCY_RANGE)
    
    # Set y-axis to start at 0
    ax.set_ylim(bottom=0)
    
    # Add legend
    ax.legend(loc='upper right', fontsize=10, framealpha=0.95, 
             edgecolor='gray', fancybox=True, shadow=False)
    ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
    
    # Tight layout
    plt.tight_layout()
    
    # Create output folder if it doesn't exist
    output_path = Path(OUTPUT_FOLDER)
    output_path.mkdir(parents=True, exist_ok=True)
    
    # Save the figure
    output_file = output_path / f'{filename}_fft.png'
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    print(f'✓ Saved: {output_file}')
    
    # Display statistics in console
    print(f'\n--- FFT Analysis for {filename} ---')
    print(f'Sampling Rate: {SAMPLING_RATE} Hz')
    print(f'Nyquist Frequency: {SAMPLING_RATE/2} Hz')
    print(f'Frequency Resolution: {SAMPLING_RATE/len(df):.4f} Hz')
    
    for label, peaks in peak_info.items():
        print(f'\n{label} - Dominant Frequencies:')
        if peaks:
            for i, (freq, mag) in enumerate(peaks, 1):
                print(f'  {i}. {freq:.2f} Hz (Magnitude: {mag:.4f} m/s²)')
        else:
            print('  No significant peaks detected')
    print('-' * 50)
    
    # Close the plot to free memory
    plt.close()
    
    # Return peak information for summary comparison
    return peak_info




def main():
    """
    Main function to process all CSV files in the specified folder.
    """
    print('=' * 60)
    print('FFT VIBRATION DATA ANALYSIS')
    print('=' * 60)
    print(f'\nData Folder: {DATA_FOLDER}')
    print(f'Output Folder: {OUTPUT_FOLDER}\n')
    
    # Find all CSV files in the data folder
    csv_files = glob.glob(str(Path(DATA_FOLDER) / '*.csv'))
    
    if not csv_files:
        print(f'⚠ No CSV files found in {DATA_FOLDER}')
        return
    
    print(f'Found {len(csv_files)} CSV file(s)\n')
    
    # Dictionary to store peak information from all files for comparison
    all_peak_info = {}
    
    # Process each CSV file
    for i, csv_file in enumerate(csv_files, 1):
        print(f'\n[{i}/{len(csv_files)}] Processing: {Path(csv_file).name}')
        try:
            peak_info = process_csv_file(csv_file)
            # Store peak information with filename as key
            all_peak_info[Path(csv_file).stem] = peak_info
        except Exception as e:
            print(f'✗ Error processing {Path(csv_file).name}: {str(e)}')
    
    print('\n' + '=' * 60)
    print('FFT ANALYSIS COMPLETE!')
    print('=' * 60)
    print(f'\nAll FFT graphs have been saved to:\n{OUTPUT_FOLDER}\n')
    
    # Create summary comparison graphs if we have data from multiple files
    if len(all_peak_info) > 1:
        print('\nGenerating summary comparison graphs...')
        create_summary_graphs(all_peak_info)
        print('✓ Summary graphs created!\n')
    elif len(all_peak_info) == 1:
        print('\nNote: Only one file processed. Summary comparison requires multiple files.\n')


def create_excel_summary(all_peak_info, output_path):
    """
    Create an Excel file with detailed FFT statistics summary.
    
    Parameters:
    -----------
    all_peak_info : dict
        Dictionary with filename as key and peak_info dict as value
    output_path : Path
        Output directory path
    """
    # Prepare data for Excel
    data_rows = []
    
    for filename, peak_info in all_peak_info.items():
        for label in LABEL_MAPPING.values():
            if label in peak_info and peak_info[label]:
                peaks = peak_info[label]
                # Add a row for each dominant peak
                for i, (freq, mag) in enumerate(peaks, 1):
                    row = {
                        'File Name': filename,
                        'Accelerometer Location': label,
                        'Peak Number': i,
                        'Frequency (Hz)': round(freq, 2),
                        'Magnitude (m/s²)': round(mag, 6)
                    }
                    data_rows.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(data_rows)
    
    # Create Excel writer object
    excel_file = output_path / 'fft_summary_statistics.xlsx'
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # Write main data sheet
        df.to_excel(writer, sheet_name='Detailed FFT Statistics', index=False)
        
        # Create pivot tables for easier comparison
        # Pivot 1: Peak 1 Frequencies
        df_peak1 = df[df['Peak Number'] == 1]
        if not df_peak1.empty:
            pivot_freq1 = df_peak1.pivot(index='File Name', 
                                         columns='Accelerometer Location', 
                                         values='Frequency (Hz)')
            pivot_freq1.to_excel(writer, sheet_name='Peak 1 Frequencies')
        
        # Pivot 2: Peak 1 Magnitudes
        if not df_peak1.empty:
            pivot_mag1 = df_peak1.pivot(index='File Name', 
                                        columns='Accelerometer Location', 
                                        values='Magnitude (m/s²)')
            pivot_mag1.to_excel(writer, sheet_name='Peak 1 Magnitudes')
        
        # Get workbook and worksheets for formatting
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        workbook = writer.book
        
        # Format the detailed statistics sheet
        ws_detail = writer.sheets['Detailed FFT Statistics']
        
        # Set column widths
        ws_detail.column_dimensions['A'].width = 35  # File Name
        ws_detail.column_dimensions['B'].width = 25  # Accelerometer Location
        ws_detail.column_dimensions['C'].width = 15  # Peak Number
        ws_detail.column_dimensions['D'].width = 20  # Frequency
        ws_detail.column_dimensions['E'].width = 25  # Magnitude
        
        # Format headers (bold)
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in ws_detail[1]:
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws_detail.iter_rows(min_row=1, max_row=len(data_rows)+1):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Data rows
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format the pivot tables
        for sheet_name in ['Peak 1 Frequencies', 'Peak 1 Magnitudes']:
            if sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]
                
                # Set column widths
                ws.column_dimensions['A'].width = 35
                for col in ['B', 'C', 'D']:
                    ws.column_dimensions[col].width = 25
                
                # Format headers
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Format first column (file names)
                for row in ws.iter_rows(min_row=2, max_col=1):
                    for cell in row:
                        cell.font = Font(bold=True)
                
                # Add borders and center alignment
                for row in ws.iter_rows(min_row=1):
                    for cell in row:
                        cell.border = thin_border
                        if cell.row > 1 and cell.column > 1:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    print(f'✓ Saved Excel summary: {excel_file}')


def create_summary_graphs(all_peak_info):
    """
    Create bar graphs comparing dominant frequencies and their magnitudes across all files.
    
    Parameters:
    -----------
    all_peak_info : dict
        Dictionary with filename as key and peak_info dict as value
    """
    # Extract data for plotting
    filenames = list(all_peak_info.keys())
    
    # Prepare data structures for the top 3 peaks of each accelerometer
    # Structure: {label: {peak_num: [freq1, freq2, ...], ...}}
    freq_data = {label: {i: [] for i in range(1, NUM_PEAKS_TO_LABEL + 1)} 
                 for label in LABEL_MAPPING.values()}
    mag_data = {label: {i: [] for i in range(1, NUM_PEAKS_TO_LABEL + 1)} 
                for label in LABEL_MAPPING.values()}
    
    # Collect data from all files
    for filename in filenames:
        peak_info = all_peak_info[filename]
        for label in LABEL_MAPPING.values():
            if label in peak_info:
                peaks = peak_info[label]
                # Pad with None if fewer than NUM_PEAKS_TO_LABEL peaks
                for i in range(NUM_PEAKS_TO_LABEL):
                    if i < len(peaks):
                        freq_data[label][i + 1].append(peaks[i][0])
                        mag_data[label][i + 1].append(peaks[i][1])
                    else:
                        freq_data[label][i + 1].append(0)  # Use 0 for missing peaks
                        mag_data[label][i + 1].append(0)
    
    # Create figure with subplots for each accelerometer location
    fig, axes = plt.subplots(len(LABEL_MAPPING), 2, figsize=(16, 5 * len(LABEL_MAPPING)))
    
    # If only one accelerometer, make sure axes is 2D
    if len(LABEL_MAPPING) == 1:
        axes = axes.reshape(1, -1)
    
    # Setup for grouped bar chart
    x = np.arange(len(filenames))
    width = 0.25  # Width of bars
    
    # Plot for each accelerometer location
    for row_idx, (label, color) in enumerate(COLORS.items()):
        ax_freq = axes[row_idx, 0]
        ax_mag = axes[row_idx, 1]
        
        # Plot dominant frequencies
        for peak_num in range(1, NUM_PEAKS_TO_LABEL + 1):
            offset = (peak_num - 2) * width  # Center the middle bar
            frequencies = freq_data[label][peak_num]
            
            # Only plot if we have non-zero values
            if any(f > 0 for f in frequencies):
                ax_freq.bar(x + offset, frequencies, width,
                           label=f'Peak {peak_num}',
                           alpha=0.8,
                           edgecolor='black',
                           linewidth=1)
        
        ax_freq.set_xlabel('CSV File', fontsize=11, fontweight='bold')
        ax_freq.set_ylabel('Frequency (Hz)', fontsize=11, fontweight='bold')
        ax_freq.set_title(f'{label} - Dominant Frequencies', fontsize=12, fontweight='bold')
        ax_freq.set_xticks(x)
        ax_freq.set_xticklabels(filenames, rotation=45, ha='right', fontsize=9)
        ax_freq.legend(loc='upper right', fontsize=9)
        ax_freq.grid(True, alpha=0.3, linestyle='--', linewidth=0.5, axis='y')
        ax_freq.set_ylim(bottom=0)
        
        # Plot dominant magnitudes
        for peak_num in range(1, NUM_PEAKS_TO_LABEL + 1):
            offset = (peak_num - 2) * width  # Center the middle bar
            magnitudes = mag_data[label][peak_num]
            
            # Only plot if we have non-zero values
            if any(m > 0 for m in magnitudes):
                ax_mag.bar(x + offset, magnitudes, width,
                          label=f'Peak {peak_num}',
                          alpha=0.8,
                          edgecolor='black',
                          linewidth=1)
        
        ax_mag.set_xlabel('CSV File', fontsize=11, fontweight='bold')
        ax_mag.set_ylabel('Magnitude (m/s²)', fontsize=11, fontweight='bold')
        ax_mag.set_title(f'{label} - Peak Magnitudes', fontsize=12, fontweight='bold')
        ax_mag.set_xticks(x)
        ax_mag.set_xticklabels(filenames, rotation=45, ha='right', fontsize=9)
        ax_mag.legend(loc='upper right', fontsize=9)
        ax_mag.grid(True, alpha=0.3, linestyle='--', linewidth=0.5, axis='y')
        ax_mag.set_ylim(bottom=0)
    
    plt.tight_layout()
    
    # Save the summary figure
    output_path = Path(OUTPUT_FOLDER)
    summary_file = output_path / 'fft_summary_comparison.png'
    plt.savefig(summary_file, dpi=300, bbox_inches='tight')
    print(f'✓ Saved summary graph: {summary_file}')
    
    plt.close()
    
    # Create Excel table with detailed FFT statistics
    create_excel_summary(all_peak_info, output_path)


if __name__ == '__main__':
    main()