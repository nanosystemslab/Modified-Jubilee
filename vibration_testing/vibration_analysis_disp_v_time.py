"""
Vibration Data Analysis Script
Processes CSV files containing accelerometer data and generates displacement vs time graphs
with RMS values and peak displacement labels.

Data Specifications:
- Sampling Rate: 100 Hz
- Raw Input: Acceleration data in g (gravitational acceleration units, 1g = 9.81 m/s²)
- Output: Displacement in mm (converted from acceleration via double integration)

Processing Method:
- Applies 5-40 Hz band-pass Butterworth filter to isolate dominant vibration band
- Integrates x, y, z acceleration components separately (more accurate)
- Calculates displacement magnitude from integrated components
- Removes DC offset and drift at each integration step
"""

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from pathlib import Path
import glob
from scipy import integrate, signal

# Configuration
DATA_FOLDER = r"C:\Users\dvier\Documents\Thesis\Vibration Testing"
OUTPUT_FOLDER = r"C:\Users\dvier\Documents\Thesis\Vibration Testing\Output_Graphs"
SAMPLING_RATE = 100  # Hz
G_TO_MS2 = 9.81  # Conversion factor: 1g = 9.81 m/s²
M_TO_MM = 1000   # Conversion factor: 1m = 1000mm

# Band-pass filter parameters
# Based on FFT analysis: dominant vibration content is in 5-40 Hz range
BANDPASS_LOW = 5.0    # Hz - lower cutoff (removes low-frequency drift)
BANDPASS_HIGH = 40.0  # Hz - upper cutoff (removes high-frequency noise)
FILTER_ORDER = 4      # Butterworth filter order

# Label mapping for accelerometer locations
LABEL_MAPPING = {
    'MPU_Direct': 'Bed Plate',
    'MPU_Ch0': 'Vertical Extrusion',
    'MPU_Ch1': 'Tool Head'
}

# Color scheme for each accelerometer
COLORS = {
    'Bed Plate': '#2E86AB',      # Blue
    'Vertical Extrusion': '#A23B72',  # Purple
    'Tool Head': '#F18F01'        # Orange
}


def calculate_rms(data):
    """
    Calculate Root Mean Square (RMS) value of a dataset.
    
    Parameters:
    -----------
    data : array-like
        Input data array
    
    Returns:
    --------
    float : RMS value
    """
    return np.sqrt(np.mean(np.square(data)))


def apply_bandpass_filter(data, low_cutoff, high_cutoff, sampling_rate, order=4):
    """
    Apply a band-pass Butterworth filter to isolate the dominant vibration band.
    
    This removes both low-frequency drift (below low_cutoff) and high-frequency
    noise (above high_cutoff), keeping only the vibration frequencies of interest.
    
    Parameters:
    -----------
    data : array-like
        Input signal data
    low_cutoff : float
        Lower cutoff frequency in Hz
    high_cutoff : float
        Upper cutoff frequency in Hz
    sampling_rate : float
        Sampling rate in Hz
    order : int
        Filter order (default: 4)
    
    Returns:
    --------
    numpy.ndarray : Filtered signal
    """
    # Calculate normalized cutoff frequencies (0 to 1, where 1 is Nyquist frequency)
    nyquist = 0.5 * sampling_rate
    low_normal = low_cutoff / nyquist
    high_normal = high_cutoff / nyquist
    
    # Design Butterworth band-pass filter
    b, a = signal.butter(order, [low_normal, high_normal], btype='band', analog=False)
    
    # Apply filter using filtfilt for zero-phase filtering
    filtered_data = signal.filtfilt(b, a, data)
    
    return filtered_data


def acceleration_to_displacement(acceleration_g, sampling_rate):
    """
    Convert acceleration data (in g) to displacement (in mm) via double integration.
    
    NOTE: This is a simplified approach that integrates the magnitude directly.
    For more accurate results, use acceleration_components_to_displacement() which
    integrates x, y, z components separately.
    
    Parameters:
    -----------
    acceleration_g : array-like
        Acceleration data in g units
    sampling_rate : float
        Sampling rate in Hz
    
    Returns:
    --------
    numpy.ndarray : Displacement in mm
    """
    # Convert g to m/s²
    acceleration_ms2 = acceleration_g * G_TO_MS2
    
    # Remove DC offset (gravity component) before integration
    acceleration_ms2 = acceleration_ms2 - np.mean(acceleration_ms2)
    
    # Calculate time step
    dt = 1.0 / sampling_rate
    
    # First integration: acceleration to velocity (m/s)
    # Use cumulative trapezoidal integration
    velocity = integrate.cumulative_trapezoid(acceleration_ms2, dx=dt, initial=0)
    
    # Remove velocity drift (detrend)
    velocity = velocity - np.mean(velocity)
    
    # Second integration: velocity to displacement (m)
    displacement_m = integrate.cumulative_trapezoid(velocity, dx=dt, initial=0)
    
    # Remove displacement drift (detrend)
    displacement_m = displacement_m - np.mean(displacement_m)
    
    # Convert to mm
    displacement_mm = displacement_m * M_TO_MM
    
    return displacement_mm


def acceleration_components_to_displacement(acc_x_g, acc_y_g, acc_z_g, sampling_rate):
    """
    Convert acceleration components (x, y, z in g) to displacement magnitude (in mm)
    via double integration of each component separately.
    
    This is the more accurate method compared to integrating the magnitude directly.
    Includes band-pass filtering to isolate the dominant vibration band before integration.
    
    Parameters:
    -----------
    acc_x_g : array-like
        X-axis acceleration in g units
    acc_y_g : array-like
        Y-axis acceleration in g units
    acc_z_g : array-like
        Z-axis acceleration in g units
    sampling_rate : float
        Sampling rate in Hz
    
    Returns:
    --------
    numpy.ndarray : Displacement magnitude in mm
    """
    # Calculate time step
    dt = 1.0 / sampling_rate
    
    # Lists to store displacement components
    displacement_components = []
    
    # Process each axis (x, y, z)
    for acc_g in [acc_x_g, acc_y_g, acc_z_g]:
        # Convert g to m/s²
        acc_ms2 = acc_g * G_TO_MS2
        
        # Remove DC offset (gravity component)
        acc_ms2 = acc_ms2 - np.mean(acc_ms2)
        
        # Apply band-pass Butterworth filter to isolate dominant vibration band
        # This removes low-frequency drift AND high-frequency noise
        acc_ms2_filtered = apply_bandpass_filter(
            acc_ms2, BANDPASS_LOW, BANDPASS_HIGH, sampling_rate, order=FILTER_ORDER
        )
        
        # First integration: acceleration to velocity (m/s)
        velocity = integrate.cumulative_trapezoid(acc_ms2_filtered, dx=dt, initial=0)
        
        # Remove velocity drift
        velocity = velocity - np.mean(velocity)
        
        # Second integration: velocity to displacement (m)
        displacement_m = integrate.cumulative_trapezoid(velocity, dx=dt, initial=0)
        
        # Remove displacement drift
        displacement_m = displacement_m - np.mean(displacement_m)
        
        displacement_components.append(displacement_m)
    
    # Calculate displacement magnitude from components
    # displacement_magnitude = sqrt(dx² + dy² + dz²)
    disp_x, disp_y, disp_z = displacement_components
    displacement_magnitude_m = np.sqrt(disp_x**2 + disp_y**2 + disp_z**2)
    
    # Convert to mm
    displacement_magnitude_mm = displacement_magnitude_m * M_TO_MM
    
    return displacement_magnitude_mm


def find_peak_displacement(time, displacement):
    """
    Find the peak (maximum absolute) displacement and its corresponding time.
    
    Parameters:
    -----------
    time : array-like
        Time values
    displacement : array-like
        Displacement values in mm
    
    Returns:
    --------
    tuple : (peak_value, peak_time, peak_index)
    """
    abs_displacement = np.abs(displacement)
    peak_index = np.argmax(abs_displacement)
    peak_value = displacement[peak_index]
    peak_time = time[peak_index]
    
    return peak_value, peak_time, peak_index


def process_csv_file(file_path):
    """
    Process a single CSV file and generate displacement vs time plot.
    
    The CSV file contains accelerometer data with x, y, z components (in g) and 
    magnitude values (also in g) for each accelerometer location.
    This function integrates the x, y, z components separately to calculate displacement,
    which is more accurate than integrating the magnitude directly.
    
    Parameters:
    -----------
    file_path : str or Path
        Path to the CSV file
    """
    # Read the CSV file
    df = pd.read_csv(file_path)
    
    # Extract filename without extension for the plot title
    filename = Path(file_path).stem
    
    # Create figure and axis
    fig, ax = plt.subplots(figsize=(14, 8))
    
    # Time data (assuming it's in the 'timestamp' column)
    time = df['timestamp'].values
    
    # Dictionary to store RMS values and peaks
    statistics = {}
    
    # Process each accelerometer location
    for old_label, new_label in LABEL_MAPPING.items():
        # Get acceleration component columns for this accelerometer
        x_column = f'{old_label}_x'
        y_column = f'{old_label}_y'
        z_column = f'{old_label}_z'
        
        if x_column in df.columns and y_column in df.columns and z_column in df.columns:
            # Get acceleration components (in g)
            acc_x = df[x_column].values
            acc_y = df[y_column].values
            acc_z = df[z_column].values
            
            # Convert acceleration components to displacement magnitude (mm)
            # This integrates each axis separately, then calculates magnitude
            displacement_mm = acceleration_components_to_displacement(
                acc_x, acc_y, acc_z, SAMPLING_RATE
            )
            
            # Calculate RMS of displacement
            rms_value = calculate_rms(displacement_mm)
            
            # Find peak displacement
            peak_value, peak_time, peak_index = find_peak_displacement(time, displacement_mm)
            
            # Store statistics
            statistics[new_label] = {
                'rms': rms_value,
                'peak_value': peak_value,
                'peak_time': peak_time,
                'peak_index': peak_index
            }
            
            # Plot the data
            ax.plot(time, displacement_mm, 
                   label=new_label,
                   color=COLORS[new_label],
                   linewidth=1.5,
                   alpha=0.8)
            
            # Mark the peak with a marker only (no text annotation)
            ax.plot(peak_time, peak_value, 
                   marker='o', 
                   markersize=8, 
                   color=COLORS[new_label],
                   markeredgecolor='black',
                   markeredgewidth=1.5,
                   linestyle='None',
                   zorder=5)
    
    # Customize the plot
    ax.set_xlabel('Time (s)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Displacement (mm)', fontsize=12, fontweight='bold')
    ax.set_title(f'Vibration Analysis: {filename}\nDisplacement vs Time (Sampling: {SAMPLING_RATE} Hz, BPF: {BANDPASS_LOW}-{BANDPASS_HIGH} Hz)', 
                fontsize=14, fontweight='bold', pad=20)
    
    # Move legend to upper left corner
    ax.legend(loc='upper left', fontsize=10, framealpha=0.95, 
             edgecolor='gray', fancybox=True, shadow=False)
    ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
    
    # Tight layout
    plt.tight_layout()
    
    # Create output folder if it doesn't exist
    output_path = Path(OUTPUT_FOLDER)
    output_path.mkdir(parents=True, exist_ok=True)
    
    # Save the figure
    output_file = output_path / f'{filename}_analysis.png'
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    print(f'✓ Saved: {output_file}')
    
    # Display statistics in console
    print(f'\n--- Statistics for {filename} ---')
    print(f'Sampling Rate: {SAMPLING_RATE} Hz')
    print(f'Duration: {time[-1]:.3f} s ({len(time)} samples)')
    for label, stats in statistics.items():
        print(f'\n{label}:')
        print(f'  RMS: {stats["rms"]:.4f} mm')
        print(f'  Peak: {stats["peak_value"]:.4f} mm at {stats["peak_time"]:.3f} s')
    print('-' * 50)
    
    # Close the plot to free memory
    plt.close()
    
    # Return statistics for summary comparison
    return statistics


def main():
    """
    Main function to process all CSV files in the specified folder.
    """
    print('=' * 60)
    print('VIBRATION DATA ANALYSIS')
    print('=' * 60)
    print(f'\nData Folder: {DATA_FOLDER}')
    print(f'Output Folder: {OUTPUT_FOLDER}\n')
    
    # Find all CSV files in the data folder
    csv_files = glob.glob(str(Path(DATA_FOLDER) / '*.csv'))
    
    if not csv_files:
        print(f'⚠ No CSV files found in {DATA_FOLDER}')
        return
    
    print(f'Found {len(csv_files)} CSV file(s)\n')
    
    # Dictionary to store statistics from all files for comparison
    all_statistics = {}
    
    # Process each CSV file
    for i, csv_file in enumerate(csv_files, 1):
        print(f'\n[{i}/{len(csv_files)}] Processing: {Path(csv_file).name}')
        try:
            stats = process_csv_file(csv_file)
            # Store statistics with filename as key
            all_statistics[Path(csv_file).stem] = stats
        except Exception as e:
            print(f'✗ Error processing {Path(csv_file).name}: {str(e)}')
    
    print('\n' + '=' * 60)
    print('ANALYSIS COMPLETE!')
    print('=' * 60)
    print(f'\nAll graphs have been saved to:\n{OUTPUT_FOLDER}\n')
    
    # Create summary comparison graphs if we have data from multiple files
    if len(all_statistics) > 1:
        print('\nGenerating summary comparison graphs and table...')
        create_summary_graphs(all_statistics)
        print('✓ Summary graphs and Excel table created!\n')
    elif len(all_statistics) == 1:
        print('\nNote: Only one file processed. Summary comparison requires multiple files.\n')


def create_excel_summary(all_statistics, output_path):
    """
    Create an Excel file with detailed statistics summary.
    
    Parameters:
    -----------
    all_statistics : dict
        Dictionary with filename as key and statistics dict as value
    output_path : Path
        Output directory path
    """
    # Prepare data for Excel
    data_rows = []
    
    for filename, stats in all_statistics.items():
        for label in LABEL_MAPPING.values():
            if label in stats:
                row = {
                    'File Name': filename,
                    'Accelerometer Location': label,
                    'RMS Displacement (mm)': round(stats[label]['rms'], 6),
                    'Peak Displacement (mm)': round(abs(stats[label]['peak_value']), 6),
                    'Peak Time (s)': round(stats[label]['peak_time'], 3)
                }
                data_rows.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(data_rows)
    
    # Create Excel writer object
    excel_file = output_path / 'summary_statistics.xlsx'
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # Write main data sheet
        df.to_excel(writer, sheet_name='Detailed Statistics', index=False)
        
        # Create pivot tables for easier comparison
        # Pivot 1: RMS values
        pivot_rms = df.pivot(index='File Name', 
                             columns='Accelerometer Location', 
                             values='RMS Displacement (mm)')
        pivot_rms.to_excel(writer, sheet_name='RMS Comparison')
        
        # Pivot 2: Peak values
        pivot_peak = df.pivot(index='File Name', 
                              columns='Accelerometer Location', 
                              values='Peak Displacement (mm)')
        pivot_peak.to_excel(writer, sheet_name='Peak Comparison')
        
        # Get workbook and worksheets for formatting
        workbook = writer.book
        
        # Format the detailed statistics sheet
        ws_detail = writer.sheets['Detailed Statistics']
        
        # Set column widths
        ws_detail.column_dimensions['A'].width = 35  # File Name
        ws_detail.column_dimensions['B'].width = 25  # Accelerometer Location
        ws_detail.column_dimensions['C'].width = 25  # RMS
        ws_detail.column_dimensions['D'].width = 25  # Peak
        ws_detail.column_dimensions['E'].width = 20  # Peak Time
        
        # Format headers (bold)
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in ws_detail[1]:
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws_detail.iter_rows(min_row=1, max_row=len(data_rows)+1):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Data rows
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format the pivot tables
        for sheet_name in ['RMS Comparison', 'Peak Comparison']:
            ws = writer.sheets[sheet_name]
            
            # Set column widths
            ws.column_dimensions['A'].width = 35
            for col in ['B', 'C', 'D']:
                ws.column_dimensions[col].width = 25
            
            # Format headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Format first column (file names)
            for row in ws.iter_rows(min_row=2, max_col=1):
                for cell in row:
                    cell.font = Font(bold=True)
            
            # Add borders and center alignment
            for row in ws.iter_rows(min_row=1):
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1 and cell.column > 1:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    print(f'✓ Saved Excel summary: {excel_file}')


def create_summary_graphs(all_statistics):
    """
    Create bar graphs comparing RMS values and peak displacements across all files.
    Also creates an Excel table with detailed statistics.
    
    Parameters:
    -----------
    all_statistics : dict
        Dictionary with filename as key and statistics dict as value
    """
    # Extract data for plotting
    filenames = list(all_statistics.keys())
    
    # Prepare data structures
    rms_data = {label: [] for label in LABEL_MAPPING.values()}
    peak_data = {label: [] for label in LABEL_MAPPING.values()}
    
    for filename in filenames:
        stats = all_statistics[filename]
        for label in LABEL_MAPPING.values():
            if label in stats:
                rms_data[label].append(stats[label]['rms'])
                peak_data[label].append(abs(stats[label]['peak_value']))
    
    # Create figure with two subplots
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))
    
    # Setup for grouped bar chart
    x = np.arange(len(filenames))
    width = 0.25  # Width of bars
    
    # Plot RMS comparison
    for i, (label, color) in enumerate(COLORS.items()):
        offset = (i - 1) * width
        ax1.bar(x + offset, rms_data[label], width, 
               label=label, color=color, alpha=0.8, 
               edgecolor='black', linewidth=1)
    
    ax1.set_xlabel('CSV File', fontsize=12, fontweight='bold')
    ax1.set_ylabel('RMS Displacement (mm)', fontsize=12, fontweight='bold')
    ax1.set_title('RMS Displacement Comparison', fontsize=14, fontweight='bold')
    ax1.set_xticks(x)
    ax1.set_xticklabels(filenames, rotation=45, ha='right', fontsize=9)
    ax1.legend(loc='upper right', fontsize=10, framealpha=0.95)
    ax1.grid(True, alpha=0.3, linestyle='--', linewidth=0.5, axis='y')
    
    # Plot Peak displacement comparison
    for i, (label, color) in enumerate(COLORS.items()):
        offset = (i - 1) * width
        ax2.bar(x + offset, peak_data[label], width, 
               label=label, color=color, alpha=0.8, 
               edgecolor='black', linewidth=1)
    
    ax2.set_xlabel('CSV File', fontsize=12, fontweight='bold')
    ax2.set_ylabel('Peak Displacement (mm)', fontsize=12, fontweight='bold')
    ax2.set_title('Peak Displacement Comparison', fontsize=14, fontweight='bold')
    ax2.set_xticks(x)
    ax2.set_xticklabels(filenames, rotation=45, ha='right', fontsize=9)
    ax2.legend(loc='upper right', fontsize=10, framealpha=0.95)
    ax2.grid(True, alpha=0.3, linestyle='--', linewidth=0.5, axis='y')
    
    plt.tight_layout()
    
    # Save the summary figure
    output_path = Path(OUTPUT_FOLDER)
    summary_file = output_path / 'summary_comparison.png'
    plt.savefig(summary_file, dpi=300, bbox_inches='tight')
    print(f'✓ Saved summary graph: {summary_file}')
    
    plt.close()
    
    # Create Excel table with detailed statistics
    create_excel_summary(all_statistics, output_path)


if __name__ == '__main__':
    main()